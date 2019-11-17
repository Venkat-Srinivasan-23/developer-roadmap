import face_recognition
import cv2
import openpyxl
import numpy as np
import datetime
import sys

# This is a demo of running face recognition on live video from your webcam. It's a little more complicated than the
# other example, but it includes some basic performance tweaks to make things run a lot faster:
#   1. Process each video frame at 1/4 resolution (though still display it at full resolution)
#   2. Only detect faces in every other frame of video.

# PLEASE NOTE: This example requires OpenCV (the `cv2` library) to be installed only to read from your webcam.
# OpenCV is *not* required to use the face_recognition library. It's only required if you want to run this
# specific demo. If you have trouble installing it, try any of the other demos that don't require it instead.

# Get a reference to webcam #0 (the default one)
video_capture = cv2.VideoCapture(0)

# Load a sample picture and learn how to recognize it.
a_image = face_recognition.load_image_file("students/1.jpeg")
a_face_encoding = face_recognition.face_encodings(a_image)[0]

b_image = face_recognition.load_image_file("students/2.jpeg")
b_face_encoding = face_recognition.face_encodings(b_image)[0]

c_image = face_recognition.load_image_file("students/3.jpeg")
c_face_encoding = face_recognition.face_encodings(c_image)[0]

d_image = face_recognition.load_image_file("students/4.jpeg")
d_face_encoding = face_recognition.face_encodings(d_image)[0]

e_image = face_recognition.load_image_file("students/5.jpeg")
e_face_encoding = face_recognition.face_encodings(e_image)[0]

f_image = face_recognition.load_image_file("students/6.jpeg")
f_face_encoding = face_recognition.face_encodings(f_image)[0]

g_image = face_recognition.load_image_file("students/7.jpeg")
g_face_encoding = face_recognition.face_encodings(g_image)[0]

h_image = face_recognition.load_image_file("students/8.jpeg")
h_face_encoding = face_recognition.face_encodings(h_image)[0]

i_image = face_recognition.load_image_file("students/9.jpeg")
i_face_encoding = face_recognition.face_encodings(i_image)[0]

# Create arrays of known face encodings and their names
known_face_encodings = [
    a_face_encoding,
    b_face_encoding,
    c_face_encoding,
    d_face_encoding,
    e_face_encoding,
    f_face_encoding,
    g_face_encoding,
    h_face_encoding,
    i_face_encoding,
]
known_face_names = [
    "Venkat",
    "Preeti",
    "Manoj",
    "Sreedhar",
    "Praveen",
    "Bala",
    "Selva",
    "Shangavi",
    "Ram"
]

# Initialize some variables
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True

workbook = openpyxl.Workbook()
sheet = workbook.active

header_name= sheet.cell(row=1,column=1)
header_time= sheet.cell(row=1,column=2)
header_hour=sheet.cell(row=1,column=3)
header_hour_name=sheet.cell(row=1,column=4)
header_name.value="Student name"
header_time.value="Entry time"
header_hour.value="Hour"
header_hour_name.value="Hour name"

row_count=1;

added_names=[]

mon=["OST","LINUX","MPC","ST","MINI PROJECT","MINI PROJECT","MINI PROJECT"]
tue=["MPC","OST","LINUX","CC","P&T","P&T","P&T"]
wed=["MPC","ST","CC","LINUX","ST LAB","ST LAB","ST LAB"]
thu=["CC","LIB","MPC","OST","LINUX","ST","P&T"]
fri=["ST","OST LAB","OST LAB","OST LAB","P&T","OST","CC"]
sat=["SEM","MINI PROJECT","MINI PROJECT","MINI PROJECT","SPD","COUN","P&T"]


while True:
    # Grab a single frame of video
    ret, frame = video_capture.read()

    # Resize frame of video to 1/4 size for faster face recognition processing
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

    # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
    rgb_small_frame = small_frame[:, :, ::-1]

    # Only process every other frame of video to save time
    if process_this_frame:
        # Find all the faces and face encodings in the current frame of video
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        face_names = []
        for face_encoding in face_encodings:
            # See if the face is a match for the known face(s)
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"

            #If a match was found in known_face_encodings, just use the first one.
            #if True in matches:
            #    first_match_index = matches.index(True)
            #    name = known_face_names[first_match_index]         

            # Or instead, use the known face with the smallest distance to the new face
            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            best_match_index = np.argmin(face_distances)
            if matches[best_match_index]:
                name = known_face_names[best_match_index]

                if not name in added_names:

                    added_names.append(name)

                    print(name)

                    row_count=row_count+1;

                    time=datetime.datetime.now()

                    name_col = sheet.cell(row = row_count, column = 1)
                    time_col = sheet.cell(row = row_count, column = 2)
                    hour_col = sheet.cell(row=row_count, column=3)
                    name_col.value=name
                    time_col.value=str(time)

                    if (int(time.hour) == 8 and int(time.minute) >= 45) or (int(time.hour) == 9 and int(time.minute) < 35):
                        hour_col.value=1
                    elif (int(time.hour) == 9 and int(time.minute) >= 35) or (int(time.hour) == 10 and int(time.minute) < 45):
                        hour_col.value=2
                    elif (int(time.hour) == 10 and int(time.minute) >= 45) or (int(time.hour) == 11 and int(time.minute) < 35):
                        hour_col.value=3
                    elif (int(time.hour) == 11 and int(time.minute) >= 35) or (int(time.hour) == 13 and int(time.minute) < 30):
                        hour_col.value=4
                    elif (int(time.hour) == 13 and int(time.minute) >= 30) or (int(time.hour) == 14 and int(time.minute) < 20):
                        hour_col.value=5
                    elif (int(time.hour) == 14 and int(time.minute) >= 20) or (int(time.hour) == 15 and int(time.minute) < 25):
                        hour_col.value=6
                    elif int(time.hour) == 15 and int(time.minute) >= 25:
                        hour_col.value=7
                    else:
                        hour_col.value="Unknown"

                    hour_name_col=sheet.cell(row=row_count,column=4)

                    if not hour_col.value == "Unknown":

                        hour = int(hour_col.value) - 1

                        #Get day of week.
                        day = time.strftime("%A")

                        if day == "Monday":
                            hour_name_col.value=mon[hour]

                        elif day == "Tuesday":
                            hour_name_col.value=tue[hour]

                        elif day == "Wednesday":
                            hour_name_col.value=wed[hour]

                        elif day == "Thursday":
                            hour_name_col.value=thu[hour]

                        elif day == "Friday":
                            hour_name_col.value=fri[hour]

                        elif day == "Saturday":
                            hour_name_col.value=sat[hour]
                   
                    else:
                        hour_name_col.value="Unknown"
        

            face_names.append(name)

    process_this_frame = not process_this_frame


    # Display the results
    for (top, right, bottom, left), name in zip(face_locations, face_names):
        
        # Scale back up face locations since the frame we detected in was scaled to 1/4 size
        top *= 4
        right *= 4
        bottom *= 4
        left *= 4

        # Draw a box around the face
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

        # Draw a label with a name below the face
        cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
        font = cv2.FONT_HERSHEY_DUPLEX
        cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)
        

    # Display the resulting image
    cv2.imshow('Video', frame)

    # Hit 'q' on the keyboard to quit!
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release handle to the webcam
workbook.save(sys.argv[1])
video_capture.release()
cv2.destroyAllWindows()
